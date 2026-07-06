"""
Populate the monthly reporting template.

Takes the user's reporting-template workbook plus converted CREFC data for one or
more deals, and writes the data into the four core data tabs, keyed by
transaction ID. Everything else in the workbook (the Dashboard and its formulas,
the Transaction List, the Dlq/REO surveillance tabs, header-file tabs, styles,
merged cells) is left untouched.

Merge behaviour: for each data tab, existing rows whose transaction ID matches an
uploaded deal are replaced; rows for deals that were not uploaded are kept. So a
monthly refresh can update just the deals that were re-downloaded.
"""

from __future__ import annotations

import io
import datetime as _dt

from openpyxl import load_workbook

from crefc_convert import date_column_indices

# CREFC file type -> reporting-template data tab
CORE_TABS = {
    "CCOL": "CMBS - Coll Summ Data",
    "LPER": "CMBS - Periodic Data",
    "PROP": "CMBS - Property Data",
    "CBND": "CMBS - Bond Data",
}

DATE_FMT = "yyyy-mm-dd"


def _txids_in_frame(df):
    col_a = df.iloc[:, 0]
    return {str(v).strip() for v in col_a if v is not None and str(v).strip()}


def build_template(template_bytes: bytes, results_by_type: dict, progress=None) -> tuple[bytes, list]:
    """Write converted data into the template's data tabs.

    results_by_type: {file_type: [DataFrame, ...]} across all uploaded deals.
    Returns (workbook_bytes, log_messages).
    """
    wb = load_workbook(io.BytesIO(template_bytes))
    log: list[str] = []

    for ft, tab in CORE_TABS.items():
        frames = results_by_type.get(ft) or []
        if not frames:
            log.append(f"{tab}: no {ft} data uploaded — left unchanged.")
            continue
        if tab not in wb.sheetnames:
            log.append(f"{tab}: tab not found in template — skipped.")
            continue

        ws = wb[tab]
        ncols = ws.max_column
        date_idx = set(date_column_indices(ft))

        uploaded_txids = set()
        for df in frames:
            uploaded_txids |= _txids_in_frame(df)

        # Keep existing rows for deals that were NOT re-uploaded.
        kept = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            a = row[0]
            if a is None or str(a).strip() == "":
                continue
            if str(a).strip() in uploaded_txids:
                continue
            kept.append(list(row))

        # New rows from the converted frames.
        new_rows = []
        for df in frames:
            for rec in df.itertuples(index=False, name=None):
                new_rows.append(list(rec))

        all_rows = kept + new_rows
        old_max = ws.max_row

        # Write data starting at row 2.
        for i, rec in enumerate(all_rows):
            r = i + 2
            for j in range(ncols):
                val = rec[j] if j < len(rec) else None
                if isinstance(val, float) and val != val:  # NaN
                    val = None
                cell = ws.cell(row=r, column=j + 1, value=val)
                if j in date_idx and isinstance(val, (_dt.date, _dt.datetime)):
                    cell.number_format = DATE_FMT

        # Clear any leftover rows from a previously larger dataset.
        last = len(all_rows) + 1
        if old_max > last:
            ws.delete_rows(last + 1, old_max - last)

        log.append(f"{tab}: {len(new_rows):,} new rows for {len(uploaded_txids)} deal(s); "
                   f"{len(kept):,} rows kept for other deals.")

    # The app can't run Excel/LibreOffice, so tell Excel to recompute every
    # formula the moment the file is opened.
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), log
