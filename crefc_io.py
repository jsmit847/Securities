"""Excel output helpers for converted CREFC tables."""

from __future__ import annotations

import io
import datetime as _dt

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
DATE_FMT = "yyyy-mm-dd"


def _excel_safe(v):
    """Coerce any pandas/numpy value into something openpyxl can write.

    Handles the cases the old float-only guard missed: pandas NA, NaT,
    numpy scalar types (np.int64, np.float64, np.bool_), and pandas
    Timestamp. Anything blank/missing becomes an empty string.
    """
    if v is None:
        return ""
    try:
        if pd.isna(v):            # NaN, NaT, pd.NA (scalar)
            return ""
    except (TypeError, ValueError):
        pass                      # arrays / unhashables: fall through
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    if isinstance(v, np.generic):  # np.int64 / np.float64 / np.bool_ / ...
        return v.item()
    return v


def _style_sheet(ws, df: pd.DataFrame):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=False)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # date formatting + reasonable column widths
    for j, col in enumerate(df.columns, start=1):
        letter = get_column_letter(j)
        series = df.iloc[:, j - 1]
        is_date = bool(series.map(lambda v: isinstance(v, (_dt.date, _dt.datetime))).any())
        if is_date:
            for cell in ws[letter][1:]:
                cell.number_format = DATE_FMT
        width = min(max(len(str(col)) + 2, 12), 42)
        ws.column_dimensions[letter].width = width


def write_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame):
    ws = wb.create_sheet(title=sheet_name[:31])
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append([_excel_safe(v) for v in r])
    _style_sheet(ws, df)
    return ws


def single_workbook_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, sheet_name, df)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def combined_workbook_bytes(results: dict) -> bytes:
    """results: {sheet_name: DataFrame}. Writes one tab per table."""
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, df in results.items():
        write_sheet(wb, sheet_name, df)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
